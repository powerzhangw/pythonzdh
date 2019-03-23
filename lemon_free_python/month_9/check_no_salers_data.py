from openpyxl import  load_workbook

#针对腾讯后台的数据 去检查我们自己后台的数据 是否一致
wb_2=load_workbook("sales.xlsx")
our_sheet=wb_2['t1']
our_data=[]#存储所有的后台数据
for i in range(2,our_sheet.max_row+1):
    our_sub_data=[]
    our_sub_data.append(i)#标记行数
    our_sub_data.append(our_sheet.cell(i,4).value)#标记学员QQ
    our_sub_data.append(our_sheet.cell(i,7).value)#存储金额
    our_sub_data.append(our_sheet.cell(i,1).value)#存储销售人员
    our_sub_data.append(our_sheet.cell(i,3).value)#标记是否老学员
    our_data.append(our_sub_data)

#首先核对全款数据
wb=load_workbook("tencent_info.xlsx")
tencent_sheet=wb['all']
tencent_data=[]#存储腾讯的数据
for i in range(2,tencent_sheet.max_row+1):
    tencent_sub_data=[]
    tencent_sub_data.append(i)#标记行数
    tencent_sub_data.append(tencent_sheet.cell(i,3).value)#标记学员QQ
    tencent_sub_data.append(tencent_sheet.cell(i,7).value)#存储金额
    tencent_data.append(tencent_sub_data)

#收集到两者数据后进行判断
for item in tencent_data:
    for value in our_data:
        if item[1]==value[1]:#qq和两者都相等 就可以称之为核对正常数据
            #更新腾讯的数据
            if tencent_sheet.cell(item[0],9).value==None and tencent_sheet.cell(item[0],10).value==None:
                tencent_sheet.cell(item[0],9).value=value[3]
                tencent_sheet.cell(item[0],10).value=value[4]

            if our_sheet.cell(value[0],10).value==None:
                our_sheet.cell(value[0],10).value='找到'
wb.save("tencent_info.xlsx")
wb_2.save("sales.xlsx")



