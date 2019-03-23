# -*- coding: utf-8 -*-
# @Time    : 2018/10/8 19:34
# @Author  : lemon_huahua
# @Email   : 204893985@qq.com
# @File    : check_installment_data.py


from openpyxl import  load_workbook

#针对腾讯后台的数据 去检查我们自己后台的数据 是否一致
wb_2=load_workbook("sales.xlsx")
our_sheet=wb_2['installment']
our_data=[]#存储所有的后台数据
for i in range(1,our_sheet.max_row+1):
    our_sub_data=[]
    our_sub_data.append(i)#标记行数
    our_sub_data.append(our_sheet.cell(i,4).value)#标记学员QQ
    our_sub_data.append(our_sheet.cell(i,8).value)#标记学分期期数
    our_sub_data.append(our_sheet.cell(i,7).value)#存储金额
    our_sub_data.append(our_sheet.cell(i,1).value)#存储销售人员
    our_sub_data.append(our_sheet.cell(i,3).value)#标记是否老学员
    our_data.append(our_sub_data)

#首先核对全款数据
wb=load_workbook("tencent_info.xlsx")
tencent_sheet=wb['installment']
tencent_data=[]#存储腾讯的数据
for i in range(1,tencent_sheet.max_row+1):
    tencent_sub_data=[]
    tencent_sub_data.append(i)#标记行数
    tencent_sub_data.append(tencent_sheet.cell(i,2).value)#标记学员QQ
    tencent_sub_data.append(tencent_sheet.cell(i,3).value)#标记学分期期数
    tencent_sub_data.append(tencent_sheet.cell(i,5).value)#存储金额
    tencent_data.append(tencent_sub_data)

#收集到两者数据后进行判断
for item in tencent_data:
    for value in our_data:
        if item[1]==value[1] and item[3]==value[3] and item[2]==value[2]:#qq和两者都相等 就可以称之为核对正常数据
            #更新腾讯的数据
            if tencent_sheet.cell(item[0],6).value!=None and tencent_sheet.cell(item[0],7).value!=None:
                tencent_sheet.cell(item[0],8).value='异常数据'
            else:
                tencent_sheet.cell(item[0],6).value=value[4]
                tencent_sheet.cell(item[0],7).value=value[5]

            if our_sheet.cell(value[0],10).value!=None:
                our_sheet.cell(value[0],11).value='异常数据'
            else:
                our_sheet.cell(value[0],10).value='找到'
wb.save("tencent_info.xlsx")
wb_2.save("sales.xlsx")



