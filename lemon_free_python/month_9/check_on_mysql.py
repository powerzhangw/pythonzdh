from openpyxl import load_workbook

def do_excel(file_name,sheet_name):
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]

    all_data=[]
    for i in range(1,sheet.max_row+1):
        sub_data=[]
        sub_data.append(sheet.cell(i,2).value)#收集QQ
        sub_data.append(sheet.cell(i,4).value[:11])#收集支付时间
        if sheet.cell(i,3).value=='首付':#收集期数
            sub_data.append(0)
        if sheet.cell(i,4)
sql="UPDATE divide_record SET actual_payment_time='{0}' " \
    "where term={1} and sales_record_id in " \
    "(select id from sales_record where member_id in " \
    "(select id FROM member where qq like '%{2}%'))".format(paytime,term,qq)
s='2018-10-07 15:43:13'
print(s[:11])